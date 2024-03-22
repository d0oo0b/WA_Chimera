import os
import sqlite3
from tkinter import *
from tkinter import filedialog, messagebox
import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
import traceback

window_width = 600
window_height = window_width

# 创建Tkinter窗口
root = Tk()
root.title("WA Chimera")

# 设置窗口大小和位置
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# 加载背景图片
bg_image = Image.open("Chimera.png")
bg_image = bg_image.resize((window_width, window_height), resample=Image.Resampling.LANCZOS)
bg_photo = ImageTk.PhotoImage(bg_image)

# 创建背景标签
bg_label = Label(root, image=bg_photo)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)


# 创建Label小部件
text_frame = Frame(root, bg="white")
text_frame.place(relx=0.5, rely=0.6, anchor="center", relwidth=0.7, relheight=0.09)
msg_label = Label(text_frame, text="Please import the xlsx file exported from AWS Trusted Advisor into WA Chimera using the button below.", bg="white", fg="black", wraplength=350, justify="center", height=2)
msg_label.pack(pady=10)




if os.path.exists('data.db'):
    # 如果文件存在,则删除它
    os.remove('data.db')
    print('**********Deleted data.db***********')
# 创建SQLite数据库连接
conn = sqlite3.connect('data.db')
c = conn.cursor()



# 创建文件选择框
def browse_file():
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Excel File", filetypes=(("Excel Files", "*.xlsx"),))
    if filename:
        import_excel(filename)

# 导入Excel文件到SQLite
def import_excel(filename):
    file_name = os.path.basename(filename).split('.')[0]
    table_name = "TA_all"
    
    # 删除同名表
    
    # 创建主表
    create_table_query = "CREATE TABLE IF NOT EXISTS " + table_name + " (check_index TEXT PRIMARY KEY, check_item TEXT, account_id TEXT, description TEXT, status TEXT)"
    c.execute(create_table_query)
    
    # 读取Excel文件
    xlsx = pd.ExcelFile(filename)
    
    # 循环每个sheet页
    for sheet_name in xlsx.sheet_names:
        df = xlsx.parse(sheet_name, header=None)
        
        # 提取标题和账号ID
        check_item = df.iloc[0, 0]
        account_id = df.iloc[1, 0].split(': ')[1]
        description = df.iloc[2, 0].split(': ')[1]
        status = df.iloc[3, 0]
        
        # 插入数据
        insert_query = "INSERT OR REPLACE INTO " + table_name + " (check_index, check_item, account_id, description, status) VALUES (?, ?, ?, ?, ?)"
        c.execute(insert_query, (sheet_name, check_item, account_id, description, status))

        try:
            df_detail = xlsx.parse(sheet_name, header=9)
            df_detail.to_sql(sheet_name, conn, if_exists='replace', index=False)
        except Exception as e:
            # 打印异常信息
            #print("An exception occurred:")
            print(e)
            # 打印异常的堆栈跟踪
            #print("Traceback:")
            #traceback.print_exc()

    # 创建Lens表
    table_name = 'lens'
    c.execute(f'DROP TABLE IF EXISTS {table_name}')

    with open('output.csv', 'r') as csvfile:
        reader = csv.reader(csvfile)
        columns = next(reader)  # 获取列名
        columns_str = ', '.join([f'"{col}"' for col in columns])
        c.execute(f'CREATE TABLE {table_name} ({columns_str} TEXT)')

        # 插入数据
        insert_query = f'INSERT INTO {table_name} VALUES ({",".join(["?"] * len(columns))})'
        c.executemany(insert_query, reader)

    # 提交更改
    conn.commit()

    output_excel()

    messagebox.showinfo("Success", f"'{filename}' analysis successful")



def output_excel():
    # 连接SQLite数据库
    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    # 执行SQL查询
    query = """
    SELECT a.check_index, b.[Pillar Name], b.[Question Title], b.[Choice Title], b.[Trusted Advisor Checks]
    FROM TA_all a, lens b
    WHERE b.[Trusted Advisor Checks] LIKE a.check_item||'%'
      AND a.status IN ('Status: warning', 'Status: error')
    ORDER BY b.[Pillar Name], b.[Question Title];
    """
    c.execute(query)
    results = c.fetchall()

    # 将结果转换为DataFrame
    columns = [desc[0] for desc in c.description][1:]  # 排除check_index列
    df = pd.DataFrame([row[1:] for row in results], columns=columns)

    # 创建新的Excel文件
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TA-check"

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for col in range(df.shape[1]):
        cell = worksheet.cell(row=1, column=col + 1)
        cell.value = df.columns[col]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 60
    worksheet.column_dimensions["C"].width = 60
    worksheet.column_dimensions["D"].width = 60

    # 写入数据
    for row in range(df.shape[0]):
        for col in range(df.shape[1]):
            cell = worksheet.cell(row=row + 2, column=col + 1)
            cell.value = df.iloc[row, col]

    # 循环结果集,创建新的工作表
    for check_index in [row[0] for row in results]:
        try:
            query = f"SELECT * FROM [{check_index}];"
            print("********Detail table SQL************")
            print(query)
            c.execute(query)
        except Exception as e:
            # 打印异常信息
            #print("An exception occurred:")
            print(e)
            # 打印异常的堆栈跟踪
            #print("Traceback:")
            #traceback.print_exc()
            continue

        detail_results = c.fetchall()
        detail_columns = [desc[0] for desc in c.description]
        detail_df = pd.DataFrame(detail_results, columns=detail_columns)

        worksheet = workbook.create_sheet(check_index[:29])

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        for col, column_name in enumerate(detail_columns, start=1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = column_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        detail_df = pd.DataFrame(detail_results, columns=detail_columns)


        # 自动调整列宽并写入数据
        for col in range(detail_df.shape[1]):
            column_letter = get_column_letter(col + 1)
            try:
                detail_df_numeric = detail_df.iloc[:, col].apply(pd.to_numeric, errors='coerce')
                detail_df_numeric = detail_df_numeric.dropna()
                max_value = detail_df_numeric.max()
                worksheet.column_dimensions[column_letter].width = max(len(str(max_value)) * 1.2, 20)
            except ValueError:
                worksheet.column_dimensions[column_letter].width = 20


        for row in range(detail_df.shape[0]):
            for col in range(detail_df.shape[1]):
                cell = worksheet.cell(row=row + 2, column=col + 1)
                cell.value = detail_df.iloc[row, col]


    # 保存Excel文件
    workbook.save("TA-check.xlsx")

    # 关闭数据库连接
    conn.close()

# 定义更新Label文本的函数
def update_msg_label(msg):
    msg_label.config(text=msg)

# 创建文件选择按钮
file_button = Button(root, text="Select Excel File", font=("Arial", 14), command=browse_file)
file_button.place(relx=0.5, rely=0.75, anchor="center")


root.mainloop()

# 关闭数据库连接
conn.close()
